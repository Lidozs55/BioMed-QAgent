"""excel_to_json.py — Excel 转 DataRecord JSON 列表。

输入：.xlsx 文件（第一行为表头，后续每行一个 record）
输出：符合 schemas/data_record.schema.json 的 DataRecord 列表

用 openpyxl 读取（沙箱中可 pip install openpyxl），自动类型推断
（openpyxl 已自动识别数字/布尔，本脚本仅做字符串兜底）。
列识别规则与 csv_to_json 一致（record_id / source_name 列优先）。

执行示例：
  python scripts/io/excel_to_json.py --input data.xlsx --out records.json \
      --task-id T1 --sheet "Sheet1" --source-name "user_upload"
"""
import os
import sys

_HERE = os.path.dirname(os.path.abspath(__file__))
if _HERE not in sys.path:
    sys.path.insert(0, _HERE)
from _base import setup_cli, make_record, save_json, emit_ok, emit_error, _hash8

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

_META_COLS = {"record_id", "source_name", "source_doi", "source_url",
              "source_pmid", "source_accession", "extraction_confidence"}


def _cell_value(cell):
    """读取单元格值，openpyxl 已推断数字/布尔，这里仅兜底字符串化与去空白。"""
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    return str(v).strip()


def convert(input_path, task_id, source_name, sheet):
    """读取 Excel 并转为 DataRecord 列表。"""
    if load_workbook is None:
        raise ImportError("openpyxl 未安装，请在沙箱中执行 pip install openpyxl")
    wb = load_workbook(filename=input_path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = ws.iter_rows()
    try:
        header = [str(c.value).strip() if c.value is not None else ""
                  for c in next(rows)]
    except StopIteration:
        wb.close()
        return []
    has_rid = "record_id" in header
    records = []
    for i, row in enumerate(rows):
        cells = [_cell_value(c) for c in row]
        rowmap = dict(zip(header, cells))
        # 收集 fields（排除元数据列与空值）
        fields = {k: v for k, v in rowmap.items()
                  if k and k not in _META_COLS and v != ""}
        if not fields:
            continue
        src = rowmap.get("source_name") or source_name
        rid = rowmap.get("record_id") if has_rid and rowmap.get("record_id") else None
        if not rid:
            seed = f"{input_path}:{i}:{sorted(rowmap.items())}"
            rid = f"xlsx-{_hash8(seed)}"
        rec = make_record(task_id, src, fields, input_path)
        rec["record_id"] = rid
        sref = rec["source_ref"]
        for k in ("source_doi", "source_url", "source_pmid", "source_accession"):
            if rowmap.get(k):
                sref[k] = str(rowmap[k])
        conf = rowmap.get("extraction_confidence")
        if conf:
            try:
                rec["extraction_confidence"] = float(conf)
            except (ValueError, TypeError):
                pass
        records.append(rec)
    wb.close()
    return records


def main():
    parser = setup_cli("excel_to_json", "Excel 转 DataRecord JSON 列表")
    parser.add_argument("--task-id", default="default", dest="task_id", help="任务 ID")
    parser.add_argument("--sheet", default=None, help="工作表名（默认第一个 sheet）")
    parser.add_argument("--source-name", default="excel_import", dest="source_name",
                        help="默认 source_name（当表无 source_name 列时使用）")
    args = parser.parse_args()
    try:
        records = convert(args.input, args.task_id, args.source_name, args.sheet)
        save_json(records, args.out)
        emit_ok(args.out, len(records))
    except Exception as e:
        emit_error(str(e))


if __name__ == "__main__":
    main()
