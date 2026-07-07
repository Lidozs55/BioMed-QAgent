"""Excel 导出工具（BioMed QAgent Stage 6）。

输入：cleaned DataRecord JSON + 可选 lineage.json
输出：Excel .xlsx 文件，两个 sheet：
  - "Data" sheet：与 to_csv 相同的列
  - "Lineage" sheet：provenance nodes 列表
    （node_id, operation_type, agent_name, tool_name, timestamp, status,
      parameters JSON, output_data_ids）

特性：
- 用 openpyxl 写入（不引入 pandas）
- 表头加粗、冻结首行、自动列宽（按内容长度简单估算）
- lineage 缺失时只生成 Data sheet

模块导入示例：
    from .export.to_excel import write_excel
    rows, sheets = write_excel(records, "data.xlsx", lineage=lineage)
"""
import json
import logging
import os

from ._base import build_columns, flatten_record, load_lineage, load_records

logger = logging.getLogger(__name__)

# Lineage sheet 固定列
LINEAGE_COLUMNS = [
    "node_id", "operation_type", "agent_name", "tool_name",
    "timestamp", "status", "parameters", "output_data_ids",
]


def _estimate_width(text):
    """按字符宽度估算列宽（中文按 2，ASCII 按 1）。"""
    if text is None:
        return 0
    s = str(text)
    w = 0
    for ch in s:
        w += 2 if ord(ch) > 127 else 1
    return w


def _auto_width(ws, columns, rows, max_width=60, min_width=8):
    """按表头与行内容估算每列宽度，应用到工作表。"""
    for col_idx, col_name in enumerate(columns, start=1):
        best = _estimate_width(col_name)
        for row in rows:
            val = row.get(col_name, "")
            best = max(best, _estimate_width(val))
        width = max(min_width, min(best + 2, max_width))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width


def _write_data_sheet(wb, records):
    """写入 Data sheet。返回行数。"""
    from openpyxl.styles import Font
    ws = wb.create_sheet("Data")
    columns = build_columns(records)
    # 表头
    ws.append(columns)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    ws.freeze_panes = "A2"
    # 数据行
    flat_rows = []
    for r in records:
        flat = flatten_record(r)
        # 确保每列都有值（缺失填空串）
        row = {c: flat.get(c, "") for c in columns}
        flat_rows.append(row)
        ws.append([row[c] for c in columns])
    _auto_width(ws, columns, flat_rows)
    return len(records), columns


def _write_lineage_sheet(wb, lineage):
    """写入 Lineage sheet。返回节点数。"""
    from openpyxl.styles import Font
    ws = wb.create_sheet("Lineage")
    ws.append(LINEAGE_COLUMNS)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    ws.freeze_panes = "A2"
    nodes = lineage.get("nodes", []) if lineage else []
    flat_rows = []
    for n in nodes:
        params = n.get("parameters", {}) or {}
        out_ids = n.get("output_data_ids", []) or []
        row = {
            "node_id": n.get("node_id", "") or "",
            "operation_type": n.get("operation_type", "") or "",
            "agent_name": n.get("agent_name", "") or "",
            "tool_name": n.get("tool_name", "") or "",
            "timestamp": n.get("timestamp", "") or "",
            "status": n.get("status", "") or "",
            "parameters": json.dumps(params, ensure_ascii=False) if params else "",
            "output_data_ids": ";".join(str(x) for x in out_ids),
        }
        flat_rows.append(row)
        ws.append([row[c] for c in LINEAGE_COLUMNS])
    _auto_width(ws, LINEAGE_COLUMNS, flat_rows)
    return len(nodes)


def write_excel(records, output_path, lineage=None):
    """写入 Excel 文件，返回 (rows, sheets)。"""
    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise RuntimeError(
            "openpyxl 未安装，请在沙箱中执行: pip install openpyxl"
        ) from e
    out_dir = os.path.dirname(os.path.abspath(output_path))
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb = Workbook()
    # 移除默认 Sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    rows, cols = _write_data_sheet(wb, records)
    sheets = ["Data"]
    if lineage:
        node_count = _write_lineage_sheet(wb, lineage)
        sheets.append("Lineage")
        logger.warning("Lineage sheet 写入 %s 个节点", node_count)
    wb.save(output_path)
    return rows, sheets
