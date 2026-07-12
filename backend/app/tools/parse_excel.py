"""Excel 文件解析器。

使用 openpyxl 解析 .xlsx / .xls 文件，将每个 Sheet 转换为
ParsedDataset 对象。依赖 openpyxl 库。

对应数据解析工具集 — 表格数据处理。
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from app.domain.processing import ParsedDataset
from app.tools.processing import _infer_field_types

PARSER_NAME = "excel_parser"
PARSER_VERSION = "0.1.0"


# ---------------------------------------------------------------------------
# Excel 解析
# ---------------------------------------------------------------------------


def parse_excel(
    file_path: str,
    dataset_id: str | None = None,
    sheet_name: str | None = None,
) -> list[ParsedDataset]:
    """解析 Excel 文件 (.xlsx/.xls)，每个 Sheet 生成一个 ParsedDataset。

    第一行作为列头，后续行为数据行。使用只读模式和数据值模式打开
    工作簿以提高性能和内存效率。

    Args:
        file_path: Excel 文件路径。
        dataset_id: 数据集 ID 前缀。None 时使用文件名。每个 Sheet 的
            dataset_id 格式为 ``{prefix}_{sheet_name}``。
        sheet_name: 指定要解析的 Sheet 名称。None 时解析所有 Sheet。

    Returns:
        ParsedDataset 列表，每个 Sheet 对应一个。

    Raises:
        ImportError: openpyxl 未安装时。
        FileNotFoundError: 文件不存在时。
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "解析 Excel 文件需要 openpyxl 库。请执行: pip install openpyxl"
        )

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")

    ds_prefix = dataset_id or path.stem
    datasets: list[ParsedDataset] = []

    # 只读模式 + 仅数据值：更快、更低内存
    workbook = openpyxl.load_workbook(
        str(path),
        read_only=True,
        data_only=True,
    )

    try:
        sheets_to_parse: list[str] = (
            [sheet_name] if sheet_name else workbook.sheetnames
        )

        for sname in sheets_to_parse:
            if sheet_name and sname not in workbook.sheetnames:
                raise ValueError(
                    f"Sheet '{sheet_name}' 不存在。可用 Sheet: {workbook.sheetnames}"
                )

            ws = workbook[sname]
            rows_raw: list[list[Any]] = [
                [cell for cell in row] for row in ws.iter_rows(values_only=True)
            ]

            if not rows_raw:
                # 空 Sheet
                datasets.append(ParsedDataset(
                    dataset_id=f"{ds_prefix}_{sname}",
                    source_file=str(path),
                    table_name=sname,
                    field_names=[],
                    field_types={},
                    rows=[],
                    source_locator=f"{sname}!A1",
                    parser_name=PARSER_NAME,
                    parser_version=PARSER_VERSION,
                ))
                continue

            # 第一行为列头
            header_raw = rows_raw[0]
            header: list[str] = [
                str(h).strip() if h is not None else f"column_{i}"
                for i, h in enumerate(header_raw)
            ]

            # 数据行
            data_rows = rows_raw[1:]
            rows: list[dict[str, Any]] = []
            for row_vals in data_rows:
                row_dict: dict[str, Any] = {}
                for j, col_name in enumerate(header):
                    row_dict[col_name] = row_vals[j] if j < len(row_vals) else ""
                rows.append(row_dict)

            field_types = _infer_field_types(header, rows)

            datasets.append(ParsedDataset(
                dataset_id=f"{ds_prefix}_{sname}",
                source_file=str(path),
                table_name=sname,
                field_names=list(header),
                field_types=field_types,
                rows=rows,
                source_locator=f"{sname}!A1",
                parser_name=PARSER_NAME,
                parser_version=PARSER_VERSION,
            ))

    finally:
        workbook.close()

    return datasets
