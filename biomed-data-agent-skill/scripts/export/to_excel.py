"""Excel 导出脚本（BioMed QAgent Stage 6）。

输入：cleaned DataRecord JSON + 可选 lineage.json
输出：Excel .xlsx 文件，两个 sheet：
  - "Data" sheet：与 to_csv 相同的列
  - "Lineage" sheet：provenance nodes 列表
    （node_id, operation_type, agent_name, tool_name, timestamp, status,
      parameters JSON, output_data_ids）

特性：
- 用 openpyxl 写入（不引入 pandas）
- 表头加粗、冻结首行、自动列宽（按内容长度简单估算）
- lineage 缺失时只生成 Data sheet

接口：
    python scripts/export/to_excel.py --input cleaned.json \
        --lineage lineage.json --out data.xlsx

成功输出（stdout）：{"status":"ok","output":"data.xlsx","rows":N}
失败输出：{"status":"error","message":"..."}
"""
import json
import os
import sys
from pathlib import Path

_HERE = os.path.dirname(os.path.abspath(__file__))
if _HERE not in sys.path:
    sys.path.insert(0, _HERE)
from _base import (  # noqa: E402
    build_columns, emit_error, emit_ok, flatten_record,
    load_lineage, load_records, log_stderr,
)

# Lineage sheet 固定列
LINEAGE_COLUMNS = [
    "node_id", "operation_type", "agent_name", "tool_name",
    "timestamp", "status", "parameters", "output_data_ids",
]


def _estimate_width(text):
    """按字符宽度估算列宽（中文按 2，ASCII 按 1）。"""
    if text is None:
        return 0
    s = str(text)
    w = 0
    for ch in s:
        w += 2 if ord(ch) > 127 else 1
    return w


def _auto_width(ws, columns, rows, max_width=60, min_width=8):
    """按表头与行内容估算每列宽度，应用到工作表。"""
    for col_idx, col_name in enumerate(columns, start=1):
        best = _estimate_width(col_name)
        for row in rows:
            val = row.get(col_name, "")
            best = max(best, _estimate_width(val))
        width = max(min_width, min(best + 2, max_width))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width


def _write_data_sheet(wb, records):
    """写入 Data sheet。返回行数。"""
    from openpyxl.styles import Font
    ws = wb.create_sheet("Data")
    columns = build_columns(records)
    # 表头
    ws.append(columns)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    ws.freeze_panes = "A2"
    # 数据行
    flat_rows = []
    for r in records:
        flat = flatten_record(r)
        # 确保每列都有值（缺失填空串）
        row = {c: flat.get(c, "") for c in columns}
        flat_rows.append(row)
        ws.append([row[c] for c in columns])
    _auto_width(ws, columns, flat_rows)
    return len(records), columns


def _write_lineage_sheet(wb, lineage):
    """写入 Lineage sheet。返回节点数。"""
    from openpyxl.styles import Font
    ws = wb.create_sheet("Lineage")
    ws.append(LINEAGE_COLUMNS)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    ws.freeze_panes = "A2"
    nodes = lineage.get("nodes", []) if lineage else []
    flat_rows = []
    for n in nodes:
        params = n.get("parameters", {}) or {}
        out_ids = n.get("output_data_ids", []) or []
        row = {
            "node_id": n.get("node_id", "") or "",
            "operation_type": n.get("operation_type", "") or "",
            "agent_name": n.get("agent_name", "") or "",
            "tool_name": n.get("tool_name", "") or "",
            "timestamp": n.get("timestamp", "") or "",
            "status": n.get("status", "") or "",
            "parameters": json.dumps(params, ensure_ascii=False) if params else "",
            "output_data_ids": ";".join(str(x) for x in out_ids),
        }
        flat_rows.append(row)
        ws.append([row[c] for c in LINEAGE_COLUMNS])
    _auto_width(ws, LINEAGE_COLUMNS, flat_rows)
    return len(nodes)


def write_excel(records, output_path, lineage=None):
    """写入 Excel 文件，返回 (rows, sheets)。"""
    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise RuntimeError(
            "openpyxl 未安装，请在沙箱中执行: pip install openpyxl"
        ) from e
    out_dir = os.path.dirname(os.path.abspath(output_path))
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb = Workbook()
    # 移除默认 Sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    rows, cols = _write_data_sheet(wb, records)
    sheets = ["Data"]
    if lineage:
        node_count = _write_lineage_sheet(wb, lineage)
        sheets.append("Lineage")
        log_stderr(f"Lineage sheet 写入 {node_count} 个节点")
    wb.save(output_path)
    return rows, sheets


def main():
    import argparse
    from _base import setup_cli
    parser = argparse.ArgumentParser(
        prog="to_excel",
        description="导出 cleaned DataRecord 列表为 Excel（Data + Lineage 双 sheet）",
    )
    parser.add_argument("--input", required=True,
                        help="输入 cleaned DataRecord JSON 文件或目录")
    parser.add_argument("--lineage", default=None,
                        help="lineage.json 路径（可选，缺失则不生成 Lineage sheet）")
    parser.add_argument("--out", required=True, help="输出 .xlsx 文件路径")
    parser.add_argument("--task-id", default="", help="任务 ID")
    args = parser.parse_args()
    try:
        records = load_records(args.input)
        log_stderr(f"加载 {len(records)} 条记录")
        lineage = load_lineage(args.lineage)
        if lineage is None and args.lineage:
            log_stderr(f"lineage 文件未找到: {args.lineage}（仅生成 Data sheet）")
        rows, sheets = write_excel(records, args.out, lineage=lineage)
        log_stderr(f"Excel 写入 {rows} 行，sheets={sheets} → {args.out}")
        emit_ok(args.out, rows=rows, sheets=sheets)
    except Exception as e:
        emit_error(f"Excel 导出失败: {e}")


if __name__ == "__main__":
    main()
